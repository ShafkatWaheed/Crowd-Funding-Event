"""
Schedule service: CRUD + bulk create + overlap detection + Excel export.
"""
from collections import defaultdict
from datetime import date, time
from io import BytesIO

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError, ForbiddenError, ConflictError
from app.models.schedule import EventScheduleItem
from app.models.event import Event
from app.models.user import User
from app.schemas.schedule import (
    ScheduleItemCreate,
    ScheduleItemUpdate,
    ScheduleItemResponse,
    ScheduleDayGroup,
)
from app.services import event as event_service


async def _get_or_404(db: AsyncSession, item_id: int) -> EventScheduleItem:
    q = select(EventScheduleItem).where(EventScheduleItem.id == item_id)
    row = (await db.execute(q)).scalar_one_or_none()
    if not row:
        raise NotFoundError("ScheduleItem", item_id)
    return row


def _parse_time(s: str) -> time:
    parts = s.split(":")
    return time(int(parts[0]), int(parts[1]))


def _times_overlap(s1: time, e1: time, s2: time, e2: time) -> bool:
    return s1 < e2 and s2 < e1


def _item_to_response(item: EventScheduleItem, overlaps: bool = False) -> ScheduleItemResponse:
    return ScheduleItemResponse(
        id=item.id,
        event_id=item.event_id,
        date=item.date.isoformat(),
        start_time=item.start_time.strftime("%H:%M"),
        end_time=item.end_time.strftime("%H:%M"),
        title=item.title,
        description=item.description,
        sort_order=item.sort_order,
        overlaps=overlaps,
        created_at=item.created_at,
    )


def _compute_overlaps(items: list[EventScheduleItem]) -> dict[int, bool]:
    """Return {item_id: True} for items whose time range overlaps with another same-date item."""
    by_date: dict[date, list[EventScheduleItem]] = defaultdict(list)
    for it in items:
        by_date[it.date].append(it)

    result: dict[int, bool] = {}
    for day_items in by_date.values():
        for i, a in enumerate(day_items):
            for b in day_items[i + 1:]:
                if _times_overlap(a.start_time, a.end_time, b.start_time, b.end_time):
                    result[a.id] = True
                    result[b.id] = True
    return result


async def list_schedule(db: AsyncSession, event_id: int) -> list[ScheduleDayGroup]:
    await event_service.get_or_404(db, event_id)
    q = (
        select(EventScheduleItem)
        .where(EventScheduleItem.event_id == event_id)
        .order_by(EventScheduleItem.date, EventScheduleItem.start_time, EventScheduleItem.sort_order)
    )
    rows = list((await db.execute(q)).scalars().all())
    overlap_map = _compute_overlaps(rows)

    by_date: dict[str, list[ScheduleItemResponse]] = defaultdict(list)
    for item in rows:
        by_date[item.date.isoformat()].append(
            _item_to_response(item, overlaps=overlap_map.get(item.id, False))
        )

    return [
        ScheduleDayGroup(date=d, items=items)
        for d, items in sorted(by_date.items())
    ]


def _validate_times(start_str: str, end_str: str) -> tuple[time, time]:
    s = _parse_time(start_str)
    e = _parse_time(end_str)
    if e <= s:
        raise ConflictError("end_time must be after start_time")
    return s, e


def _validate_date_in_range(d_str: str, event: Event) -> date:
    d = date.fromisoformat(d_str)
    if event.start_time and d < event.start_time.date():
        raise ConflictError("Schedule date is before event start date")
    if event.end_time and d > event.end_time.date():
        raise ConflictError("Schedule date is after event end date")
    return d


async def _check_organizer(db: AsyncSession, event: Event, user: User) -> None:
    if not await event_service.user_can_edit_event(db, event, user):
        raise ForbiddenError("Only the organizer can manage the schedule")


async def create_item(
    db: AsyncSession, event_id: int, data: ScheduleItemCreate, user: User
) -> ScheduleItemResponse:
    event = await event_service.get_or_404(db, event_id)
    await _check_organizer(db, event, user)
    s, e = _validate_times(data.start_time, data.end_time)
    d = _validate_date_in_range(data.date, event)

    item = EventScheduleItem(
        event_id=event_id,
        date=d,
        start_time=s,
        end_time=e,
        title=data.title,
        description=data.description,
        sort_order=data.sort_order,
    )
    db.add(item)
    await db.flush()
    await db.refresh(item)
    return _item_to_response(item)


async def update_item(
    db: AsyncSession, item_id: int, data: ScheduleItemUpdate, user: User
) -> ScheduleItemResponse:
    item = await _get_or_404(db, item_id)
    event = await event_service.get_or_404(db, item.event_id)
    await _check_organizer(db, event, user)

    update_data = data.model_dump(exclude_unset=True)

    new_start_str = update_data.get("start_time", item.start_time.strftime("%H:%M"))
    new_end_str = update_data.get("end_time", item.end_time.strftime("%H:%M"))
    s, e = _validate_times(new_start_str, new_end_str)

    if "date" in update_data:
        _validate_date_in_range(update_data["date"], event)
        item.date = date.fromisoformat(update_data["date"])
    item.start_time = s
    item.end_time = e

    for field in ("title", "description", "sort_order"):
        if field in update_data:
            setattr(item, field, update_data[field])

    await db.flush()
    await db.refresh(item)
    return _item_to_response(item)


async def delete_item(db: AsyncSession, item_id: int, user: User) -> None:
    item = await _get_or_404(db, item_id)
    event = await event_service.get_or_404(db, item.event_id)
    await _check_organizer(db, event, user)
    await db.delete(item)
    await db.flush()


async def bulk_create(
    db: AsyncSession, event_id: int, items: list[ScheduleItemCreate], user: User
) -> list[ScheduleItemResponse]:
    event = await event_service.get_or_404(db, event_id)
    await _check_organizer(db, event, user)

    created = []
    for data in items:
        s, e = _validate_times(data.start_time, data.end_time)
        d = _validate_date_in_range(data.date, event)
        item = EventScheduleItem(
            event_id=event_id,
            date=d,
            start_time=s,
            end_time=e,
            title=data.title,
            description=data.description,
            sort_order=data.sort_order,
        )
        db.add(item)
        created.append(item)

    await db.flush()
    for it in created:
        await db.refresh(it)
    return [_item_to_response(it) for it in created]


async def export_schedule_xlsx(db: AsyncSession, event_id: int) -> BytesIO:
    """Generate an Excel workbook with one sheet per date."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    groups = await list_schedule(db, event_id)
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    for group in groups:
        ws = wb.create_sheet(title=group.date)
        ws.append(["Time", "Title", "Description"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for item in group.items:
            ws.append([
                f"{item.start_time} – {item.end_time}",
                item.title,
                item.description or "",
            ])

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 50

    if not groups:
        ws = wb.create_sheet(title="Schedule")
        ws.append(["No schedule items yet."])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
